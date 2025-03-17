import os
import pandas as pd
import pymysql
from sqlalchemy import create_engine
import openpyxl
from openpyxl.styles import Alignment, PatternFill

# 连接 MySQL 数据库
db_config = {
    'host': '47.251.73.120',
    'user': 'joker',
    'password': 'Iamajoker666!',
    'database': 'keyword_db'
}
engine = create_engine(f"mysql+pymysql://{db_config['user']}:{db_config['password']}@{db_config['host']}/{db_config['database']}")

# Excel 文件路径
excel_path = "keyword_data_filtered_by_package.xlsx"

# 定义词包对应的背景颜色（RGB 格式）
package_colors = {
    'apoem': 'C6EFCE',  # 浅绿色
    'APoEM-私护词包': 'E6D7F0',  # 浅紫色
    # 如果有更多词包，可以继续添加颜色，例如：
    # '另一个词包': 'FFEB9C',  # 浅黄色
}

# 1. **查询 MySQL 获取最新数据**
query = "SELECT date, keyword, monthpv, package FROM keyword_data ORDER BY package, keyword;"
df = pd.read_sql(query, engine)

# 2. **按词包分组，并为每个词包生成透视表**
# 获取所有词包
packages = df['package'].unique()

# 创建一个空的 DataFrame 用于合并结果
df_final = pd.DataFrame()

# 按词包分组处理
for package in packages:
    # 筛选当前词包的数据
    df_package = df[df['package'] == package].copy()
    
    # 透视数据，使 keyword 变为列名，仅保留 monthpv
    df_pivot = df_package.pivot(index="date", columns="keyword", values="monthpv").reset_index()
    
    # 为列名添加词包前缀（可选，便于区分）
    df_pivot.columns = ['date'] + [f"{keyword}" for keyword in df_pivot.columns[1:]]
    
    # 如果是第一个词包，直接赋值
    if df_final.empty:
        df_final = df_pivot
    else:
        # 按 date 合并，保留所有列
        df_final = pd.merge(df_final, df_pivot, on='date', how='outer')

# 3. **填充 NaN 值**
df_final = df_final.fillna(0)  # 可以根据需求改为 None 或其他值

# 4. **检查 Excel 是否存在**
if os.path.exists(excel_path):
    # **如果存在，则读取原始 Excel**
    df_existing = pd.read_excel(excel_path)
    
    # **合并数据，按 date 合并**
    df_combined = pd.merge(df_existing, df_final, on='date', how='outer')
    df_combined = df_combined.fillna(0)  # 填充合并后的 NaN 值
else:
    # **如果 Excel 不存在，直接使用新数据**
    df_combined = df_final

# 5. **导出数据到 Excel**
df_combined.to_excel(excel_path, index=False)

# 6. **使用 openpyxl 自动调整列宽、居中对齐，并设置背景颜色**
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# 获取列名和对应的词包（需要重新查询数据库以获取 keyword 和 package 的映射）
keyword_to_package = df.groupby('keyword')['package'].first().to_dict()

# 为表头设置背景颜色
header_row = ws[1]  # 第一行是表头
for cell in header_row:
    keyword = cell.value
    if keyword == 'date':  # date 列不设置颜色
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # 灰色
        continue
    package = keyword_to_package.get(keyword, None)
    if package and package in package_colors:
        color = package_colors[package]
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

# 遍历所有列，调整宽度并居中对齐
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # 获取列字母
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
            # 设置单元格内容居中
            cell.alignment = Alignment(horizontal='center', vertical='center')
        except:
            pass
    adjusted_width = (max_length + 16)  # 增加一些空间
    ws.column_dimensions[column].width = adjusted_width

# 为数据区域设置背景颜色（从第二行开始）
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        col_idx = cell.column - 1  # 列索引（从 0 开始）
        keyword = ws[1][col_idx].value  # 获取表头（第一行）的 keyword
        if keyword == 'date':  # date 列不设置颜色
            continue
        package = keyword_to_package.get(keyword, None)
        if package and package in package_colors:
            color = package_colors[package]
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

# 保存调整后的 Excel 文件
wb.save(excel_path)

print(f"Excel 文件已更新，按词包分组，设置背景颜色，列宽已自动调整，内容已居中：{excel_path}")